import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from pandas.tseries.offsets import CustomBusinessDay
import requests
from datetime import datetime
from os import getlogin
from requests import post

# Função para enviar dados de log para o banco de dados Firebase


def enviardados(erro, modulo="Desconhecido"):
    """Função para registrar logs no Firebase."""
    dados = {
        "Automação": "Atualizar data Ateste",
        "Data Fim": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "Data Inicio": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "Erro": erro,
        "Modulo": modulo,
        "valores": getlogin(),
        "Versao": "V0"
    }
    try:
        response = post(
            "https://automacoesppi-default-rtdb.firebaseio.com/Execucoes.json",
            json=dados)
        print(f"Status Code: {response.status_code}")
        print(f"Response: {response.text}")
    except Exception as e:
        print(f"Erro ao enviar dados para o Firebase: {e}")

# Obter feriados nacionais via API


def obter_feriados_api(anos):
    feriados = []
    for ano in anos:
        url = f"https://brasilapi.com.br/api/feriados/v1/{ano}"
        response = requests.get(url)
        if response.status_code == 200:
            feriados.extend([feriado['date'] for feriado in response.json()])
        else:
            print(f"Erro ao obter feriados para {ano}: {response.status_code}")
    return feriados

# Adicionar feriados específicos de São Paulo


def obter_feriados_sao_paulo(anos):
    feriados_sp = []
    for ano in anos:
        feriados_sp.extend([
            # Adicione outros feriados municipais de São Paulo, se necessário
            f"{ano}-01-25",  # Aniversário de São Paulo
            f"{ano}-07-09",
        ])
    return feriados_sp

# Defina o intervalo de anos para o qual os feriados serão buscados


anos = range(2020, 2030)  # Exemplo: 2020 a 2029

# Obter feriados nacionais e regionais
feriados_nacionais = obter_feriados_api(anos)
feriados_sao_paulo = obter_feriados_sao_paulo(anos)

# Combinar feriados nacionais e regionais
feriados_completos = pd.to_datetime(list(set(
    feriados_nacionais + feriados_sao_paulo)))

# Criar o calendário customizado com feriados combinados
dias_uteis = CustomBusinessDay(holidays=feriados_completos)

# Funções auxiliares


def tratar_apf(apf):
    """Trata e normaliza o número da APF"""
    return str(apf).replace(".", "").replace("-", "").lstrip("0")


def encontrar_data_proxima(dados, coluna_data):
    """Encontra a data mais próxima para o mesmo APF"""
    dados_validos = dados[dados[coluna_data].notna()]
    if not dados_validos.empty:
        return dados_validos[coluna_data].min()
    return None


def calcular_data_util(data_vencimento):
    """Calcula a data 4 dias úteis antes da data de vencimento"""
    if pd.isna(data_vencimento):
        return None
    return (data_vencimento - 4 * dias_uteis).date()


def carregar_planilhas(arquivo_principal, arquivo_referencia):
    """Carrega as planilhas de dados"""
    df_principal = pd.read_excel(arquivo_principal)
    df_referencia = pd.read_excel(
        # arquivo_referencia, sheet_name="Acompanhamento Ateste")
        arquivo_referencia, sheet_name="Planilha1")
    return df_principal, df_referencia


def atualizar_planilha_referencia(
        df_principal, df_referencia, arquivo_referencia):
    """Atualiza a planilha de referência com os dados processados"""
    wb = load_workbook(arquivo_referencia)
    # ws = wb["Acompanhamento Ateste"]
    ws = wb["Planilha1"]

    # Definir o título em negrito na coluna 20
    ws.cell(row=1, column=20).font = Font(bold=True)
    ws.cell(row=1, column=20).value = "Ateste Atualizado"

    verde_negrito = Font(bold=True, color="00FF00")  # Verde
    vermelho_negrito = Font(bold=True, color="FF0000")  # Vermelho

    for row_num, row in enumerate(ws.iter_rows(
            min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column),
              2):
        apf_referencia = row[2].value
        if not apf_referencia:
            continue

        apf_tratado = tratar_apf(apf_referencia)
        vencimento_ateste_antigo = row[11].value

        dados_principal = df_principal[df_principal['APF'].apply(
                                        tratar_apf) == apf_tratado]
        if not dados_principal.empty:
            validade_ateste = dados_principal.iloc[0]["VALIDADE DO ATESTE"]
            junta_comercial = dados_principal.iloc[0]["JUNTA COMERCIAL"]
            if pd.isna(junta_comercial):
                junta_comercial = encontrar_data_proxima(dados_principal,
                                                         "JUNTA COMERCIAL")
            cnd_iptu = dados_principal.iloc[0][
                "CND IPTU – CERTIDÃO NEGATIVA DE TRIBUTOS IMÓVEL"]
            matricula = dados_principal.iloc[0]["MATRÍCULA DO IMÓVEL"]

            # Garantir que as datas sejam convertidas corretamente
            if pd.notna(validade_ateste):
                validade_ateste = pd.to_datetime(
                    validade_ateste, errors='coerce', dayfirst=True)
                ws.cell(row=row_num, column=12).value = validade_ateste
                ws.cell(row=row_num, column=12).number_format = "DD/MM/YYYY"
            if pd.notna(junta_comercial):
                junta_comercial = pd.to_datetime(
                    junta_comercial, errors='coerce', dayfirst=True)
                ws.cell(row=row_num, column=8).value = junta_comercial
                ws.cell(row=row_num, column=8).number_format = "DD/MM/YYYY"
            ws.cell(row=row_num, column=9).value = cnd_iptu
            ws.cell(row=row_num, column=7).value = matricula

            # Conversão para datetime, caso sejam strings
            vencimento_ateste_antigo = pd.to_datetime(
                vencimento_ateste_antigo, errors='coerce', dayfirst=True
                ) if isinstance(
                    vencimento_ateste_antigo, str
                    ) else vencimento_ateste_antigo

            # Verifique se ambas as variáveis são válidas antes de comparar
            if pd.notna(validade_ateste) and pd.notna(
                    vencimento_ateste_antigo):
                if validade_ateste.date() == vencimento_ateste_antigo.date():
                    ws.cell(row=row_num, column=20).value = "Não"
                    # "Não" em vermelho
                    ws.cell(row=row_num, column=20).font = vermelho_negrito
                else:
                    # "Sim" em verde
                    ws.cell(row=row_num, column=20).value = "Sim"
                    ws.cell(row=row_num, column=20).font = verde_negrito

            nova_data_vencimento = calcular_data_util(validade_ateste)

            if nova_data_vencimento:
                ws.cell(row=row_num, column=10).value = f"ABRIR {
                    nova_data_vencimento.strftime('%d/%m')}"

    wb.save(arquivo_referencia)
